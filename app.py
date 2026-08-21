import chainlit as cl
import requests
import pandas as pd
import json
import urllib3
import os
from dotenv import load_dotenv
from database import get_db
from models import Question, OnboardingSession
from database import engine, SessionLocal
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


import asyncio
from functools import partial
					
from auth import verify_password, load_users

load_dotenv()
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ==========================================
# SECTION 1: Azure OpenAI Configuration
# ==========================================
"""
def call_azure_llm(user_message: str, system_prompt: str = "") -> str:
    azure_url = "https://spikeiam-genai-resource.cognitiveservices.azure.com/openai/responses?api-version=2025-04-01-preview"
    api_key = os.getenv("AZURE_API_KEY") 
    
    if not api_key:
        return '{"status": "error", "message": "Error: Missing AZURE_API_KEY"}'

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    
    payload = {
        "input": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_message}
        ],
        "model": "gpt-5.4-mini"
    }
    
    try:
        response = requests.post(azure_url, headers=headers, json=payload, verify=False)
        response.raise_for_status()
        
        response_data = response.json()
        
        # 👇 ECCO LA RIGA CORRETTA CON LA NUOVA STRUTTURA DEL JSON 👇
        return response_data["output"][0]["content"][0]["text"]
        
    except Exception as e:
        error_details = str(e)
        if 'response' in locals() and response.text:
            error_details += f" | Response: {response.text}"
        return json.dumps({"status": "error", "message": f"API Error: {error_details}"})
"""


def call_azure_llm(user_message: str, system_prompt: str = "", json_mode: bool = False) -> str:
    azure_url = "https://spikeiam-genai-resource.cognitiveservices.azure.com/openai/responses?api-version=2025-04-01-preview"
    api_key = os.getenv("AZURE_API_KEY")

    if not api_key:
        return '{"status": "error", "message": "Error: Missing AZURE_API_KEY"}'

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }

    payload = {
        "input": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_message}
        ],
        "model": "gpt-5.4-mini"
    }

    if json_mode:
        payload["text"] = {"format": {"type": "json_object"}}

    try:
        response = requests.post(azure_url, headers=headers, json=payload, verify=False)
        response.raise_for_status()
        return response.json()["output"][0]["content"][0]["text"]

    except Exception as e:
        error_details = str(e)
        if 'response' in locals() and response.text:
            error_details += f" | Response: {response.text}"
        return json.dumps({"status": "error", "message": f"API Error: {error_details}"})



# ==========================================
# SECTION 2: Dynamic Excel Reading
# ==========================================
def load_questions_from_excel(file_path: str, sheet_name: str) -> list:
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        # Assuming the column is still named 'Domanda' in your Excel. 
        # Change to 'Question' if you translate the Excel header too.
        return df['Question'].dropna().tolist()
    except Exception as e:
        print(f"Error reading the Excel file: {e}")
        # Fallback questions in case of error
        return [
            "Is the target system exposed to the internet or only available on the intranet?",
            "What authentication protocol does it use?",
            "Is there a test environment separated from the production one?"
        ]
    
# ==========================================
# SECTION 3: Dynamic DB Reading
# ==========================================
def load_questions_from_DB(system_type: str) -> list:
    try:
        db = SessionLocal()

        questions_db = db.query(Question).where(Question.system_type == system_type)
        questions = []

        for r in questions_db:
            questions.append(r.question)

        db.close()

        return questions
    except Exception as e:
        print(f"Error connecting/reading DB: {e}")
        # Fallback questions in case of error
        return [
            "Is the target system exposed to the internet or only available on the intranet?",
            "What authentication protocol does it use?",
            "Is there a test environment separated from the production one?"
        ]







@cl.password_auth_callback
def auth_callback(username: str, password: str):
    users = load_users()
    print(f"=== LOGIN DEBUG ===")
    print(f"Username tentato: '{username}'")
    print(f"Utenti nel file: {list(users.keys())}")
    print(f"Username trovato: {username in users}")

    if username not in users:
        print("ERRORE: username non trovato")
        return None

    is_valid = verify_password(password, users[username])
    print(f"Password valida: {is_valid}")

    if not is_valid:
        print("ERRORE: password errata")
        return None

    print("LOGIN OK")
    return cl.User(
        identifier=username,
        metadata={"role": "user", "provider": "credentials"}
    )




# ==========================================
# SECTION 4: Initial Flow Management
# ==========================================
@cl.on_chat_start
async def start():
    # Setup session state
    cl.user_session.set("answers", {})
    cl.user_session.set("step", "company")
    
    await cl.Message(
        content="### 👋 Welcome to Spike IAM Onboarding\n\nTo configure your environment, please enter the **Company Name**:"
    ).send()

@cl.on_message
async def main(message: cl.Message):

    step = cl.user_session.get("step")
    answers = cl.user_session.get("answers")
    
    # --- STEP 1: Company ---
    if step == "company":
        cl.user_session.set("company", message.content)
        cl.user_session.set("step", "system")
        await cl.Message(
            content=f"🏢 **Company:** {message.content}\n\nGreat. What is the **Name of the Target System** we are integrating?"
        ).send()
        
    # --- STEP 2: System ---
    elif step == "system":
        cl.user_session.set("system", message.content)
        cl.user_session.set("step", "system_type")
        
        actions = [
            cl.Action(name="choose_type", payload={"value": "Others"}, label="Others"),
            cl.Action(name="choose_type", payload={"value": "AD-Azure"}, label="AD-Azure"),
            cl.Action(name="choose_type", payload={"value": "SAP"}, label="SAP"),
            cl.Action(name="choose_type", payload={"value": "LDAP"}, label="LDAP")
        ]
        
        await cl.Message(
            content=f"✅ Target System: **{message.content}**.\n\nWhat **type** of target system is it? Choose an option below to load the specific questions.",
            actions=actions
        ).send()
        
    # --- EXTRA CHECK: User types instead of clicking the button ---
    elif step == "system_type":
        await cl.Message(content="⚠️ **Please use the buttons above** to select the system type.").send() 
# --- EXTRA CHECK: User types instead of clicking the button (per la scelta del metodo) ---
    elif step == "choose_method":
        await cl.Message(content="⚠️ **Please use the buttons above** to select how you want to proceed (Chat or Excel).").send()

    # --- STEP 3A: Gestione dell'Upload Excel ---
    elif step == "upload_excel":
        if not message.elements:
            await cl.Message(content="⚠️ Please upload the completed Excel file using the attachment button (📎).").send()
            return

        file = message.elements[0]
        system_type = cl.user_session.get("system_type")
        answers = cl.user_session.get("answers") or {}
        questions = cl.user_session.get("questions") or []

        try:
            df = pd.read_excel(file.path, sheet_name=system_type)
            df.columns = df.columns.str.strip()

            if 'Answer' not in df.columns:
                await cl.Message(content="⚠️ Cannot find the column **'Answer'** in your uploaded file. Please add it, fill in your answers, and upload it again.").send()
                return

            normalized_questions = {q.strip(): q for q in questions}

             # DEBUG
            print("normalized_questions keys:", list(normalized_questions.keys())[:3])

            rows_to_validate = []
            for index, row in df.iterrows():
                q_raw = str(row.get('Question', '')).strip()
                a = row.get('Answer')

                # DEBUG
                print(f"Row {index} | q_raw: '{q_raw}' | in normalized: {q_raw in normalized_questions} | a: '{a}'")

                if q_raw not in normalized_questions:
                    continue
                if not pd.notna(a) or not str(a).strip():
                    continue
                rows_to_validate.append((q_raw, normalized_questions[q_raw], str(a).strip()))

            await cl.Message(content=f"🔄 Validating **{len(rows_to_validate)}** answers in parallel, please wait...").send()

            def validate_single(q_raw: str, answer_text: str) -> dict:
                validation_prompt = f"""You are an expert IAM technical consultant reviewing onboarding questionnaire answers.

QUESTION: "{q_raw}"
ANSWER: "{answer_text}"

Your task: decide if this answer provides ANY useful information to the question, even if incomplete or misspelled.
Accept it unless it is completely meaningless or a clear refusal.

Reply ONLY with valid JSON:
{{
    "valid": true | false,
    "reason": "One sentence explanation in English"
}}
"""
                validation_str = call_azure_llm(user_message="", system_prompt=validation_prompt, json_mode=True)
                try:
                    clean = validation_str.replace("```json", "").replace("```", "").strip()
                    result = json.loads(clean)
                    return {
                        "question": q_raw,
                        "valid": result.get("valid", False),
                        "reason": result.get("reason", "No reason provided.")
                    }
                except Exception:
                    return {
                        "question": q_raw,
                        "valid": False,
                        "reason": f"Could not parse response: {validation_str[:200]}"
                    }

            loop = asyncio.get_event_loop()
            tasks = [
                loop.run_in_executor(None, validate_single, q_raw, answer_text)
                for q_raw, original_key, answer_text in rows_to_validate
            ]
            results = await asyncio.gather(*tasks)

            extracted_count = 0
            invalid_answers = []

            for i, validation_result in enumerate(results):
                q_raw, original_key, answer_text = rows_to_validate[i]
                if validation_result["valid"]:
                    answers[original_key] = answer_text
                    extracted_count += 1
                else:
                    invalid_answers.append({
                        "question": q_raw,
                        "answer": answer_text,
                        "reason": validation_result["reason"]
                    })

            cl.user_session.set("answers", answers)

            await cl.Message(content=f"✅ **File processed successfully!** Extracted **{extracted_count}** valid answers.").send()

            if invalid_answers:
                warning_lines = ["⚠️ **The following answers were flagged as insufficient and skipped:**\n"]
                for item in invalid_answers:
                    warning_lines.append(f"- **Q:** {item['question']}\n  **A:** {item['answer']}\n  **Reason:** {item['reason']}")
                await cl.Message(content="\n\n".join(warning_lines)).send()

            cl.user_session.set("step", "conversational_chat")
            await ask_next_question(last_user_input="I have uploaded the Excel file. Please review.")

        except Exception as e:
            await cl.Message(content=f"⚠️ Error reading the file. Ensure it's a valid Excel format. Error details: {str(e)}").send()



    # --- STEP 3: Conversational Chat (Extraction, Validation, Explanation, CORRECTION) ---
    elif step == "conversational_chat":
        questions = cl.user_session.get("questions")
        
        # Recuperiamo la domanda specifica che l'LLM ha deciso di fare al giro precedente
        current_question = cl.user_session.get("current_asked_question")
        pending_questions = [q for q in questions if q not in answers or not answers[q]]
        
        if not pending_questions and not message.content:
            return # Interview already completed
            
        async with cl.Step(name="Intent and Data Analysis"):
            # UNIFIED PROMPT: Valida, Estrae multipli target, e gestisce CORREZIONI
            system_prompt_orchestrator = f"""You are an expert technical assistant in IAM (Identity and Access Management).
Your goal is to gather technical information from a user.

CURRENT QUESTION ASKED TO THE USER: "{current_question}"

REMAINING QUESTIONS TO BE SATISFIED:
{json.dumps(pending_questions, ensure_ascii=False)}

ALREADY ANSWERED QUESTIONS (Current State):
{json.dumps(answers, ensure_ascii=False)}

ANALYZE THE USER'S MESSAGE AND CHOOSE ONE OF 3 ACTIONS:
1. "clarification": The user didn't understand the question, asks "what does it mean?", or asks for help. Provide a technical explanation in "message" (in the same language the user speaks).
2. "invalid": The user tries to answer, but the response is "I don't know" or too vague to be accepted. Explain why you need more details in "message" (in the same language the user speaks).
3. "success": The user provides a valid answer AND/OR corrects a previously given answer. 
   - Extract the answer. 
   - CRITICAL RULE: Extract the answer in the EXACT SAME LANGUAGE the user wrote it (e.g., if the user answers in Italian, the extracted text MUST be in Italian). DO NOT translate it to English.
   - Map the new information to ANY relevant question in 'REMAINING QUESTIONS TO BE SATISFIED'.
   - IMPORTANT CORRECTION RULE: If the user states they made a mistake or explicitly provides updated information for a topic they already answered, map the new data to the exact question string found in 'ALREADY ANSWERED QUESTIONS'.
   - Use "message" to give a brief success feedback (in the same language the user speaks).

REPLY ONLY AND EXCLUSIVELY WITH THIS JSON:
{{
    "status": "clarification" | "invalid" | "success",
    "message": "Your response message for the user",
    "extracted_data": {{
        "EXACT text of the question (either from REMAINING or ALREADY ANSWERED array)": "Extracted, cleaned, and summarized answer in the USER'S ORIGINAL LANGUAGE"
    }}
}}
Note: "extracted_data" must be populated ONLY if status is "success"."""

            analysis_str = call_azure_llm(user_message=message.content, system_prompt=system_prompt_orchestrator)
            print(f"\n--- DEBUG RISPOSTA API ---\n{analysis_str}\n--------------------------\n")
        try:
            clean_json = analysis_str.replace("```json", "").replace("```", "").strip()
            analysis = json.loads(clean_json)
            
            status = analysis.get("status", "error")
            response_message = analysis.get("message", "Comprehension error.")
            extracted_data = analysis.get("extracted_data", {})

            if status == "clarification" or status == "invalid":
                await cl.Message(content=f"💡 {response_message}").send()
                await cl.Message(content=f"---\n**Getting back to our setup:** {current_question}").send()
                
            elif status == "success":
                # Salva sia le risposte nuove che le CORREZIONI di quelle vecchie
                for q, a in extracted_data.items():
                    if q in questions: # Controlla che la domanda esista nell'Excel
                        if q in answers:
                            # Era già stata risposta -> CORREZIONE
                            await cl.Message(content=f"🔄 *Updated requirement:* **{q}** \n> {a}").send()
                        else:
                            # È una domanda nuova -> NUOVO INSERIMENTO
                            await cl.Message(content=f"✅ *Saved requirement:* **{q}** \n> {a}").send()
                        answers[q] = a # Aggiorna o crea la chiave nel dizionario
                
                cl.user_session.set("answers", answers)
                await cl.Message(content=response_message).send()
                
                # Chiediamo la PROSSIMA domanda
                await ask_next_question(last_user_input=message.content)
                
            else:
                await cl.Message(content=f"⚠️ API Error Details: {response_message}").send()

        except json.JSONDecodeError:
            await cl.Message(content="⚠️ *The server responded in an unexpected format. Please try again.*").send()


# ==========================================
# CALLBACK: Target Type Selection
# ==========================================
@cl.action_callback("choose_type")
async def on_choose_type(action: cl.Action):
    try:
        system_type = action.payload.get("value")
        if system_type == "Others":
            system_type = await on_choose_other_target_system_type()

        cl.user_session.set("system_type", system_type)

        questions = await cl.make_async(load_questions_from_DB)(system_type)
        cl.user_session.set("questions", questions)
        cl.user_session.set("step", "choose_method")

        actions = [
            cl.Action(name="choose_method", payload={"value": "chat"}, label="💬 Continue in Chat"),
            cl.Action(name="choose_method", payload={"value": "excel"}, label="📊 Download & Upload Excel"),
        ]

        file_path = "Obiettivi AI - Target Systems.xlsx"
        df = pd.DataFrame({"Question": questions, "Answer": [""] * len(questions)})

        def build_excel():
            from openpyxl import load_workbook  # <-- import spostato qui dentro
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            df.to_excel(file_path, index=False, sheet_name=system_type[:31], engine="openpyxl")

            wb = load_workbook(file_path)
            ws = wb.active

            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_alignment = Alignment(vertical="top", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )
            alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2), start=2):
                for cell in row:
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                    if row_idx % 2 == 0:
                        cell.fill = alt_fill

            ws.column_dimensions[get_column_letter(1)].width = 60
            ws.column_dimensions[get_column_letter(2)].width = 40
            ws.freeze_panes = "A2"

            wb.save(file_path)

        await cl.make_async(build_excel)()

        await cl.Message(
            content=f"✅ System type **{system_type}** selected.\n\nHow would you like to provide the technical requirements?",
            actions=actions,
        ).send()

    except Exception as e:
        await cl.Message(content=f"❌ Errore durante la selezione del target system: `{e}`").send()
        raise

# Corrected async function: fixes early-stop bug by requiring a minimum
# number of Q&A exchanges and a "CONFIDENT" flag before stopping early.
# The evaluation LLM now returns "label|confidence" internally, but only
# "Target DB" or "Generic" are ever shown to the user or returned.

async def on_choose_other_target_system_type() -> str:
    max_questions = 7
    min_questions = 3  # Minimum number of Q&A exchanges before early-stop is allowed
    conversation_history = []

    system_prompt_ask = """You are a Senior Technical Consultant conducting a formal IAM integration assessment.
Your aim is to understand what is the target system type in order to integrate it in the IGA system.
You only know that the target system is not AD, Azure, SAP, nor LDAP; but you don't know what's the intended integration method, you have to discover it.
Keep in mind that the user doesn't know what it means to integrate a target system in an IGA system, you have to inquiry him on all the possible integration methods - APIs, DBs, ...

INSTRUCTIONS:
1. Analyze the PREVIOUS CONTEXT. Identify the main topics the user just talked about.
2. Ask the ONE question that logically follows the previous context to keep a fluid conversation.
3. Use a highly professional, polite, and formal B2B tone.
4. Be precise and clear. Do NOT use informal greetings.

Reply ONLY and EXCLUSIVELY with the question you want to ask."""

    system_prompt_evaluate = """You are an expert system architect performing a rigorous technical classification.
Carefully analyze the ENTIRE conversation below before deciding — do not rely only on the last message.
Determine whether the target system integration is a "Target DB" or "Generic".

Classification rules (apply equal rigor to both labels — never treat one as a default fallback):
- "Target DB": use this label ONLY if the conversation contains EXPLICIT and UNAMBIGUOUS evidence that the system's user/account data is managed via direct database access (e.g. explicit mention of SQL, stored procedures, direct read/write on DB tables).
- "Generic": use this label ONLY if the conversation contains EXPLICIT and UNAMBIGUOUS evidence that the integration method is something OTHER than direct database access (e.g. explicit mention of APIs, web services, connectors, flat files, or any other non-DB method).

You must ALWAYS provide your best-guess label, even if the evidence is vague or incomplete — never refuse to guess.
Additionally, provide a confidence flag:
- Reply "CONFIDENT" ONLY if there is explicit, unambiguous evidence in the conversation clearly supporting your chosen label.
- Otherwise, reply "NOT_CONFIDENT" while still providing your best-guess label (do not default to "Generic" for convenience — justify it with the same rigor as "Target DB").

Reply ONLY and EXCLUSIVELY with the two tokens separated by a single pipe character, in this exact format:
"Target DB|CONFIDENT", "Target DB|NOT_CONFIDENT", "Generic|CONFIDENT", or "Generic|NOT_CONFIDENT"."""

    async with cl.Step(name="Identifying Target System Type"):
        last_label = None  # Keeps track of the last evaluated label for best-effort fallback

        for i in range(max_questions):
            context = "\n".join(conversation_history) if conversation_history else "No previous context yet."
            user_msg_for_llm = f"PREVIOUS CONTEXT:\n{context}\n\nAsk the next question."

            question = call_azure_llm(
                user_message=user_msg_for_llm,
                system_prompt=system_prompt_ask
            )

            res = await cl.AskUserMessage(content=f"💬 {question}", timeout=300).send()

            if res is None:
                await cl.Message(content="⏱️ Timeout reached. Defaulting to 'Custom connector'.").send()
                return "Custom connector"

            user_answer = res["output"]
            conversation_history.append(f"Q: {question}")
            conversation_history.append(f"A: {user_answer}")

            eval_context = "\n".join(conversation_history)
            evaluation = call_azure_llm(
                user_message=f"CONVERSATION:\n{eval_context}",
                system_prompt=system_prompt_evaluate
            )

            # Parse the "label|confidence" response from the evaluation LLM.
            raw_evaluation = evaluation.strip()
            if "|" in raw_evaluation:
                label_part, confidence_part = raw_evaluation.split("|", 1)
                label = label_part.strip()
                confidence = confidence_part.strip().upper()
            else:
                # Defensive fallback in case the LLM doesn't respect the format.
                label = raw_evaluation.strip()
                confidence = "NOT_CONFIDENT"

            last_label = label  # Keep the most recent label for best-effort fallback at the end

            # Only consider stopping early once at least min_questions exchanges
            # have happened AND the evaluation is CONFIDENT.
            enough_exchanges = (i + 1) >= min_questions
            if enough_exchanges and confidence == "CONFIDENT":
                await cl.Message(
                    content=f"✅ Target system type identified: **{label}**"
                ).send()
                return label

        # Max questions reached without a CONFIDENT evaluation:
        # take the last evaluation's label as the best-effort final answer.
        best_effort_label = last_label if last_label else "Generic"
        await cl.Message(
            content=f"⚠️ Maximum questions reached. Best-effort target system type: **{best_effort_label}**"
        ).send()

        return best_effort_label

# ==========================================
# CALLBACK: Method Selection (Chat vs Excel)
# ==========================================
@cl.action_callback("choose_method")
async def on_choose_method(action: cl.Action):
    method = action.payload.get("value")
    
    if method == "chat":
        cl.user_session.set("step", "conversational_chat")
        await ask_next_question(last_user_input="")
        
    elif method == "excel":
        cl.user_session.set("step", "upload_excel")
        
        # Inviamo il file Excel esistente all'utente
        elements = [
            cl.File(
                name="Obiettivi AI - Target Systems.xlsx",
                path="Obiettivi AI - Target Systems.xlsx", # Il tuo file locale
                display="inline"
            )
        ]
        
        await cl.Message(
            content="📥 **Please download the Excel file attached above.**\n\n"
                    "**Instructions:**\n"
                    "1. Open the sheet corresponding to your system (**" + cl.user_session.get("system_type") + "**).\n"
                    "2. Add a new column named exactly **Answer** next to the questions.\n"
                    "3. Fill in your answers and save the file.\n\n"
                    "When you are ready, **upload the completed file here** using the attachment button (📎).",
            elements=elements
        ).send()

# ==========================================
# FUNCTION: Asks the next question dynamically
# ==========================================
async def ask_next_question(last_user_input: str = ""):
    questions = cl.user_session.get("questions")
    answers = cl.user_session.get("answers")
    
    # Calcola quali domande mancano all'appello
    pending_questions = [q for q in questions if q not in answers or not answers[q]]
    
    # --- BLOCCO 1: FINE INTERVISTA, TRADUZIONE E SALVATAGGIO DB ---
    if not pending_questions:
        translated_answers = {}
        
        async with cl.Step(name="Elaborazione e Traduzione Dati"):
            system_prompt_translate = """You are an expert IT technical translator. 
I will provide a JSON dictionary containing questions as keys and user answers as values.
Your task is to translate ALL the values (the answers) into professional IT English.
If a value is already in English, keep it exactly as it is.
CRITICAL: DO NOT translate or modify the keys (the questions).
Respond ONLY and EXCLUSIVELY with the valid translated JSON object. No markdown, no greetings."""

            # Chiamata all'LLM di Azure per la traduzione
            translation_response = call_azure_llm(
                user_message=json.dumps(answers, ensure_ascii=False), 
                system_prompt=system_prompt_translate
            )
            
            try:
                clean_json = translation_response.replace("```json", "").replace("```", "").strip()
                translated_answers = json.loads(clean_json)
            except Exception as e:
                print(f"Errore durante la traduzione: {e}")
                translated_answers = {"error": "Traduzione fallita", "raw_response": translation_response}

        # Prepariamo i dati dalla sessione per il salvataggio
        company_name = cl.user_session.get("company")
        target_system_name = cl.user_session.get("system")
        system_type_name = cl.user_session.get("system_type")

        print("\n" + "="*50 + "\n💾 [SALVATAGGIO NEL DATABASE IN CORSO...]\n" + "="*50)
        
        try:
            # Apriamo la connessione al DB e salviamo il record
            with get_db() as db:
                nuova_sessione = OnboardingSession(
                    company=company_name,
                    target_system=target_system_name,
                    system_type=system_type_name,
                    collected_data_original=answers,
                    collected_data_english=translated_answers
                )
                db.add(nuova_sessione)
                db.commit()
                
                print(f"✅ Dati salvati con successo per la company: {company_name}")
                
        except Exception as e:
            print(f"❌ Errore critico durante il salvataggio nel DB: {e}")

        # Messaggio finale all'utente
        await cl.Message(
            content="🎉 **Interview completed.** We have successfully gathered all the necessary technical requirements.\n\nThe data has been securely saved to our system. Thank you for your time."
        ).send()
        
        return # Termina l'esecuzione della funzione qui

    # --- BLOCCO 2: SELEZIONE E INVIO DELLA PROSSIMA DOMANDA ---
    
    # Creiamo un dizionario numerato per evitare che l'LLM sbagli a copiare le stringhe
    numbered_pending = {str(i): q for i, q in enumerate(pending_questions)}

    async with cl.Step(name="Contextual Question Selection"):
        system_prompt_ask = f"""You are a Senior Technical Consultant conducting a formal IAM integration assessment.

PREVIOUS CONTEXT / LAST USER MESSAGE:
"{last_user_input if last_user_input else 'None. Start of the technical interview.'}"

REMAINING QUESTIONS TO ASK (Numbered Dictionary):
{json.dumps(numbered_pending, ensure_ascii=False, indent=2)}

INSTRUCTIONS:
1. Analyze the PREVIOUS CONTEXT. Identify the main topics the user just talked about (e.g., AD, environments, users, groups, licenses, provisioning).
2. Look at the REMAINING QUESTIONS TO ASK. Select the ONE question that logically and semantically follows the PREVIOUS CONTEXT to keep a fluid conversation. 
3. If there is no clear connection, or if it's the start of the interview, always select the question with index "0".
4. Rephrase the selected question in a highly professional, polite, and formal B2B tone.
5. Be precise and clear. Do NOT use informal greetings.

Reply ONLY and EXCLUSIVELY with valid JSON in this format:
{{
    "selected_target_index": "The string key of the chosen question from the dictionary (e.g., '0', '3', '5')",
    "conversational_question": "Your rephrased, professional B2B question"
}}"""

        # Chiamata all'LLM di Azure per selezionare la prossima domanda
        response_str = call_azure_llm(user_message="", system_prompt=system_prompt_ask)
        
        try:
            clean_json = response_str.replace("```json", "").replace("```", "").strip()
            data = json.loads(clean_json)
            
            # Recuperiamo la stringa originale della domanda usando l'ID scelto dall'LLM
            selected_index = str(data.get("selected_target_index", "0"))
            
            if selected_index in numbered_pending:
                target_question = numbered_pending[selected_index]
            else:
                target_question = pending_questions[0] # Fallback di sicurezza
                
            conversational_question = data.get("conversational_question", f"Could you please provide information regarding this requirement: {target_question}")
            
        except json.JSONDecodeError:
            target_question = pending_questions[0]
            conversational_question = f"Could you please elaborate on the following requirement: {target_question}"

    # Salva in sessione la domanda esatta che stiamo per fare
    cl.user_session.set("current_asked_question", target_question)
    
    # Invia la domanda all'utente
    await cl.Message(content=f"💬 {conversational_question}").send()

